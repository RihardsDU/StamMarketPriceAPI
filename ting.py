import os
import requests
import openpyxl
import time
import ast
import configparser


# Set up default params (overwrite in init_config)
url = "https://steamcommunity.com/market/priceoverview/"
file_name = "scripExcel.xlsx"
currency = "9"
delay = 1
offset = 0
my_items = {"appid":[
    "item1",
    "item2"
]}


def get_item_data(app_id, item):
    """Get pricing data for a given item."""
    params = {
        "appid": app_id,
        "market_hash_name": item,
        "currency": currency
    }
    response = requests.get(url, params=params)
    if response.status_code == requests.codes.ok:
        try:
            return response.json()
        except ValueError:
            # Invalid JSON response
            pass
    # Return default values if response is not valid or empty
    return {
        "success":False,
        "lowest_price": "-",
        "volume": "-",
        "median_price": "-"
    }


def get_all_item_data():
    """Get pricing data for all items in the config."""
    print("\n*** GETTING ITEMS ***\n")
    item_data = {}
    count = 0
    for app_id, items in my_items.items():
        for item in items:
            if count >= offset:
                item_data[item] = get_item_data(app_id, item)
                if item_data[item]["success"]:
                    print("Got: " + item)
                else:
                    print("Failed: " + item)
                time.sleep(delay)
            count += 1
    return item_data


def check_if_file_exists():
    """Check if the excel document exists"""
    if not os.path.exists(file_name):
        print("Could not locate file " + file_name)
        print ("Creating new excel document " + file_name)
        workbook = openpyxl.Workbook()
        workbook.save(file_name)


def save_to_excel(item_data):
    """Save pricing data to an Excel file."""
    print("\n*** SAVING ITEMS ***\n")
    if not os.path.exists(file_name):
        print(f"Error: Could not locate file {file_name}")
        return
    try:
        workbook = openpyxl.load_workbook(file_name)
    except Exception as e:
        print(f"Error: Could not load workbook {file_name}: {str(e)}")
        return

    worksheet = workbook.active
    count = offset + 2
    for item, data in item_data.items():
        if count >= offset + 2:
            worksheet.cell(row=count, column=1, value=item)
            try:
                worksheet.cell(row=count, column=2, value=data["lowest_price"])
            except (KeyError, TypeError):
                worksheet.cell(row=count, column=2, value="-")
            try:
                worksheet.cell(row=count, column=3, value=data["volume"])
            except (KeyError, TypeError):
                worksheet.cell(row=count, column=3, value="-")
            try:
                worksheet.cell(row=count, column=4, value=data["median_price"])
            except (KeyError, TypeError):
                worksheet.cell(row=count, column=4, value="-")
            print(f"Added: {item}")
        count += 1
        

    try:
        workbook.save(file_name)
    except Exception as e:
        print(f"Error: Could not save workbook {file_name}: {str(e)}")


def init_config():
    config = configparser.ConfigParser()
    config.read('config.ini', encoding='utf-8')

    global url 
    global file_name
    global currency
    global delay
    global offset
    global my_items

    try:
        url = config.get('API', 'url')
        file_name = config.get('API', 'file_name')
        currency = config.get('API', 'currency')
        delay = int(config.get('API', 'delay'))
        offset = int(config.get('API', 'offset'))
        my_items_str = config.get('API', 'my_items')
        my_items = ast.literal_eval(my_items_str)
    except configparser.Error as e:
        print(f"Error reading configuration file: {e}")
        input("Press ENTER to exit")
        exit(1)
    except ValueError as e:
        print(f"Error parsing configuration value: {e}")
        input("Press ENTER to exit")
        exit(1)


def main():
    init_config()
    check_if_file_exists()
    item_data = get_all_item_data()
    save_to_excel(item_data)
    input("Press ENTER to exit")


if __name__ == "__main__":
    main()